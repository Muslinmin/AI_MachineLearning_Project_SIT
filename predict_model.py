import pandas as pd
import joblib
from sklearn.metrics import classification_report, mean_squared_error, mean_absolute_error
import argparse
import json
import os
import numpy as np
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from preprocessing import preprocess_data

def apply_color_coding(excel_file):
    # Open the Excel file and access the worksheet
    wb = load_workbook(excel_file)
    ws = wb.active

    # Define color fills
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply color to 'Survived?' and 'Insure?' columns based on values
    for row in ws.iter_rows(min_row=2, min_col=ws.max_column - 1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if cell.value == "Yes":
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    # Apply color to 'Survived?' column as well
    for row in ws.iter_rows(min_row=2, min_col=ws.max_column - 2, max_col=ws.max_column - 2, max_row=ws.max_row):
        for cell in row:
            if cell.value == "Yes":
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    # Save the colored workbook
    wb.save(excel_file)
    print(f"Color-coding applied and saved to {excel_file}")

def predict(test_data, model_path, config_file, metrics_file):
    # Load the test data
    data = pd.read_csv(test_data)
    X, y = preprocess_data(data)

    # Load hyperparameters from config
    with open(config_file) as f:
        config = json.load(f)
    model_params = config['model_params']

    # Load the trained model
    model = joblib.load(model_path)
    y_pred = model.predict(X)

    # Map predictions to human-readable labels
    target_names = ['Predicted Not Survive', 'Predicted Survive']
    y_pred_labels = np.where(y_pred == 1, 'Yes', 'No')

    # Add predictions back to the original dataframe
    data['Survived?'] = np.where(y == 1, 'Yes', 'No')
    data['Insure?'] = y_pred_labels

    # Generate the output file name dynamically
    base_filename = os.path.basename(test_data)
    file_name, file_ext = os.path.splitext(base_filename)
    output_excel = f"{file_name}_prediction_output.xlsx"

    # Save updated dataframe with predictions to a new Excel file
    data.to_excel(output_excel, index=False)
    print(f"Predictions saved to {output_excel}")

    # Apply color coding to the output Excel file
    apply_color_coding(output_excel)

    # Generate and print classification report
    test_report = classification_report(y, y_pred, target_names=target_names)
    print("Test set report:\n", test_report)

    # Write the report to a text file with a dynamic name
    report_filename = f"{file_name}_classification_report.txt"
    with open(report_filename, "w") as f:
        f.write(test_report)

    # Calculate MSE, RMSE, and MAE
    mse = mean_squared_error(y, y_pred)
    rmse = float(np.sqrt(mse))
    mae = mean_absolute_error(y, y_pred)

    print(f"Test set MSE: {mse:.4f}, RMSE: {rmse:.4f}, MAE: {mae:.4f}")

    # Write metrics JSON
    os.makedirs(os.path.dirname(metrics_file), exist_ok=True)
    metrics = {
        "report_type": "test",
        "hyperparameters": model_params,
        "test": {
            "mse": round(float(mse), 6),
            "rmse": round(rmse, 6),
            "mae": round(float(mae), 6),
            "classification_report": test_report
        }
    }
    with open(metrics_file, "w") as f:
        json.dump(metrics, f, indent=2)
    print(f"Metrics saved to {metrics_file}")

    # Call generate_pdf.py to create a PDF report
    output_pdf = f"{file_name}_report.pdf"
    subprocess.run(["python3", "generate_pdf.py", "--report", output_pdf, "--metrics", metrics_file])

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Predict using a trained model.")
    parser.add_argument('--test', type=str, help="Path to the test dataset.")
    parser.add_argument('--model', type=str, help="Path to the trained model.")
    parser.add_argument('--config', type=str, help="Path to the configuration JSON file.")
    parser.add_argument('--metrics', type=str, help="Path to save the metrics JSON file.")

    args = parser.parse_args()
    predict(args.test, args.model, args.config, args.metrics)
